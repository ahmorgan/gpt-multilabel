from openpyxl import load_workbook
import csv


def process_data(file_name) -> None:
    # access proper excel file for preprocessing
    text_workbook = load_workbook(filename="raw_data/gpt-test.xlsx")
    dataset_workbook = load_workbook(filename="raw_data/gpt-zeroes.xlsx")

    text = text_workbook.active
    dataset = dataset_workbook.active

    # excluded less common labels for now
    labels = {
        "None": 1,
        "Python and Coding": 2,
        "Github": 3,
        "MySQL": 4,
        "Assignments ": 5,  # space after "Assignments" is not a typo;
        # funnily enough, in the original dataset's "Issue" dropdown,
        # there is a stray space after the issues "Assignments", "Quizzes",
        # and "Understanding requirements and instructions". It is just
        # easier to fix it this way.
        "Quizzes ": 6,
        "Understanding requirements and instructions ": 7,
        "Learning New Material": 8,
        "Course Structure and Materials": 9,
        "Time Management and Motivation": 10,
        "Group Work": 11,
        "API": 12,
        "Project": 13
    }

    response_num = 1

    current_response = text.cell(row=1, column=2).value

    # iterate over zeroes and text spreadsheets concurrently and populate as needed
    # zeroes excel file will become the dataset we use
    for row in text.iter_rows(min_row=1, min_col=1, max_row=text.max_row, max_col=2, values_only=True):
        if current_response != row[1]:
            current_response = row[1]
            response_num += 1
        try:
            dataset.cell(row=response_num, column=labels[row[0]]).value = "1"
        except KeyError:
            print("Key not found, skipping label")

    # save populated dataset as a csv file
    # credit: https://stackoverflow.com/questions/10802417/how-to-save-an-excel-worksheet-as-csv
    with open(file_name, "w", encoding="utf-8", newline="") as f:
        c = csv.writer(f)
        for r in dataset.rows:
            l = []
            for cell in r:
                if cell.value == 0.0:
                    l.append(0)
                else:
                    l.append(1)
            c.writerow(l)





